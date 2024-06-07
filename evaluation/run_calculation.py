
import os
import argparse
import json
import logging
from openpyxl import Workbook, load_workbook

from utils import get_result_dir
from models.load_model import load_model
from dataset.load_dataset import load_dataset


# valid_models = ['minigpt4_vicuna-7b', 'minigpt4_vicuna-13b', 'minigpt4_llama_2', 'minigpt_v2', 'blip2_flan-t5-xl', 'blip2-opt-3b', 'blip2-opt-7b', 'instructblip_vicuna-7b', 'instructblip_flan-t5-xl', 'instructblip_flan-t5-xxl', 'llava_1.5-7b', 'llava_1.5-13b', 'otter', 'emu2-chat']
valid_datasets = ['vlm_test', 'pope', 'amber-ds', 'amber-g', 'vlbold-g', 'vlbold-pi', 'vlbold-p', 'vlbold-r', 'vlbold-ri', 'vlbold_xl-g', 'vlbold_xl-pi', 'vlbold_xl-p', 'vlbold_xl-r', 'vlbold_xl-ri', 'mathvista', 'mmbench', 'mmbench-cn', 'sciqa', 'ccbench', 'faithscore-qa', 'faithscore-cap', 'gaive', 'haelm', 'mmhal', 'openchair', 'ptqa', 'ptqa_mini_v5', 'toxicity-b', 'toxicity-po', 'toxicity-pr', 'chef', 'mmcbench', 'typographic', 'figstep', 'mmsafetybench', 'sap', 'minivlbbq', 'minivlbbq_reason']
parser = argparse.ArgumentParser(description='Example argparse script.')
parser.add_argument('--model_list', nargs='+', help='A list of valid models.')
parser.add_argument('--dataset_list', nargs='+', choices=valid_datasets, help='A list of valid datasets.')
parser.add_argument('--time', type=str, help='Automatically select the first evaluation result after the time.')
parser.add_argument('--cfg_options', type=json.loads,
        help='override some settings in the used config, the key-value pair '
        'in xxx=yyy format will be merged into config file. If the value to '
        'be overwritten is a list, it should be like key="[a,b]" or key=a,b '
        'It also allows nested list/tuple values, e.g. key="[(a,b),(c,d)]" '
        'Note that the quotation marks are necessary and that no white space '
        'is allowed.')
parser.add_argument('--comment', type=str, default="", help='Extra flag append to the model name in output dir.')
args = parser.parse_args()



for dataset_name in args.dataset_list:
    xlsx_path = os.path.join("./outputs", f'{dataset_name}.xlsx')
    if not os.path.exists(xlsx_path):
        workbook = Workbook()
        workbook.save(xlsx_path)

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active

    dataset = load_dataset(dataset_name)
    for model_name in args.model_list:
        result_dir, result_time = get_result_dir(model_name, dataset_name, args.time)
        print(result_dir)
        result_json = json.load(open(os.path.join(result_dir, 'result.json'), 'r'))
        score, *other = dataset.calculate_result(result_json)
        with open(os.path.join(result_dir, 'score.json'), 'w') as json_file:
            json.dump(score, json_file, indent=4)
        if other:
            with open(os.path.join(result_dir, 'other.json'), 'w') as json_file:
                json.dump(other, json_file, indent=4)

        logger = logging.getLogger(model_name + "|" + dataset_name)
        logger.setLevel(logging.INFO)
        file_handler = logging.FileHandler(os.path.join(result_dir, 'log.txt'), mode='a')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(logging.Formatter(''))
        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(logging.INFO)
        stream_handler.setFormatter(logging.Formatter('%(name)s - %(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(file_handler)
        logger.addHandler(stream_handler)
        
        for key, value in score.items():
            logger.info(f'{key}:\t{value}')
            logger.info('-' * 60)
        first_row = ['model'] + list(score.keys()) + ['time']
        append_row = [model_name] + list(score.values()) + [result_time]
        for col, data in enumerate(first_row, start=1):
            sheet.cell(row=1, column=col).value = data
        sheet.append(append_row)

        stream_handler.close()
        file_handler.close()


    workbook.save(xlsx_path)